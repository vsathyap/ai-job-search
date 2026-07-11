#!/usr/bin/env python3
"""Regenerate job_search_tracker.xlsx from job_search_tracker.csv.

job_search_tracker.csv is the single source of truth (read/written by
/scrape and /outcome). This script produces a read-friendly Excel mirror
of it - run after any command that modifies the CSV.
"""
import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = REPO_ROOT / "job_search_tracker.csv"
XLSX_PATH = REPO_ROOT / "job_search_tracker.xlsx"


def main() -> None:
    if not CSV_PATH.exists():
        print(f"{CSV_PATH} does not exist - nothing to sync.", file=sys.stderr)
        sys.exit(1)

    with CSV_PATH.open(newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))

    if not rows:
        print(f"{CSV_PATH} is empty - nothing to sync.", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    header, *data_rows = rows
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in data_rows:
        ws.append(row)

    # Autosize columns (capped, so a long notes/URL field doesn't blow up the sheet)
    for i, col_name in enumerate(header, start=1):
        max_len = len(col_name)
        for row in data_rows:
            if i - 1 < len(row):
                max_len = max(max_len, len(row[i - 1]))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)

    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH} ({len(data_rows)} application rows)")


if __name__ == "__main__":
    main()
